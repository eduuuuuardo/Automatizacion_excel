import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import NamedStyle, Font
import string

def automatizacion_excel(nombre_archivo):
    """Input sales_mes.xlsx / Output report_mes.xlsx"""
    try:
        # Leer el archivo Excel
        archivo_excel = pd.read_excel(nombre_archivo)

        # Mostrar las columnas 'Gender', 'Product line' y 'Total'
        print(archivo_excel[['Gender', 'Product line', 'Total']])

        # Crear tabla pivote
        tabla_pivote = archivo_excel.pivot_table(index='Gender', columns='Product line', values='Total', aggfunc='sum').round(0)

        mes_extensiones = nombre_archivo.split('_')[1].replace('.xlsx', '')  # Asegurarse de que esté limpio

        # Guardar tabla pivote en un nuevo archivo Excel
        tabla_pivote.to_excel('sales_2021.xlsx', startrow=4, sheet_name='Report')

        # Cargar el archivo correcto donde está la hoja 'Report'
        try:
            wb = load_workbook(f'sales_2021.xlsx')  # Asegúrate de que se carga el archivo correcto
        except FileNotFoundError as e:
            print(f"Error al cargar el archivo: {e}")
            return

        # Acceder a la hoja 'Report'
        pestaña = wb['Report']

        # Obtener las dimensiones de la hoja activa
        min_col = pestaña.min_column
        max_col = pestaña.max_column
        min_fila = pestaña.min_row
        max_fila = pestaña.max_row

        # Crear gráfico de barras
        barchart = BarChart()

        # Obtener los datos para el gráfico
        data = Reference(pestaña, min_col=min_col + 1, max_col=max_col, min_row=min_fila, max_row=max_fila)
        categoria = Reference(pestaña, min_col=min_col, max_col=min_col, min_row=min_fila + 1, max_row=max_fila)

        # Agregar datos y categorías al gráfico
        barchart.add_data(data, titles_from_data=True)
        barchart.set_categories(categoria)
        pestaña.add_chart(barchart, 'B12')

        # Agregar título y estilo al gráfico
        barchart.title = 'Ventas'
        barchart.style = 2

        # Crear un estilo para la celda de moneda
        if 'Currency' not in wb.named_styles:
            currency_style = NamedStyle(name='Currency', number_format='$#,##0.00')
            wb.add_named_style(currency_style)

        # Crear lista del abecedario
        abecedario = list(string.ascii_uppercase)
        abecedario_excel = abecedario[0:max_col]

        # Sumar los valores en cada columna y dar estilo de moneda
        for i in abecedario_excel:
            if i != 'A':
                pestaña[f'{i}{max_fila + 1}'] = f'=SUM({i}{min_fila + 1}:{i}{max_fila})'
                pestaña[f'{i}{max_fila + 1}'].style = 'Currency'

        # Agregar el título "Total" en la celda correspondiente
        pestaña[f'{abecedario_excel[0]}{max_fila + 1}'] = 'Total'

        # Agregar encabezados
        pestaña['A1'] = 'Reporte'
        mes = mes_extensiones.split(',')[0]  # Ajustar esto según el formato esperado
        pestaña['A2'] = mes

        # Aplicar estilos de fuente
        pestaña['A1'].font = Font(name='Arial', bold=True, size=20)
        pestaña['A2'].font = Font(name='Arial', bold=True, size=12)

        # Guardar el archivo
        wb.save('sales_2021.xlsx')
        print("Archivo guardado correctamente como 'sales_2021.xlsx'.")

    except Exception as e:
        print(f"Ha ocurrido un error: {e}")

# Llamar a la función
automatizacion_excel('supermarket_saless.xlsx')
